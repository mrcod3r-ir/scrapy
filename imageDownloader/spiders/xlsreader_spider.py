from asyncio.windows_events import NULL
from fileinput import filename
from importlib.resources import path
import pathlib
from pickle import FALSE
from time import sleep
from turtle import width
import scrapy
# import openpyxl module
import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import os
from PIL import Image


def pathExtractor(path: str):
  baseUrls = ['http://diet2.gosnalban.com/back/storage/app/public/','https://Core.zirehapp.com']
  Path = path
  for url in baseUrls:
    if Path.find(url) > -1:
      Path = Path.lstrip(url)
    Path = Path.lstrip('/')
  return Path.split('/')

  
def download(url: str,fileName:str, dest_folder: str,isThumnail: bool):
  if not os.path.exists(dest_folder):
      os.makedirs(dest_folder)  # create folder if it does not exist

  filename = fileName or url.split('/')[-1].replace(" ", "_")  # be careful with file names
  file_path = os.path.join(dest_folder, filename)
  # ====================
  session = requests.Session()
  retry = Retry(connect=3, backoff_factor=0.5)
  adapter = HTTPAdapter(max_retries=retry)
  session.mount('http://', adapter)
  session.mount('https://', adapter)
  r = session.get(url)
  # ====================
  
  if r.ok:
      print("saving to", os.path.abspath(file_path))
      with open(file_path, 'wb') as f:
          for chunk in r.iter_content(chunk_size=1024 * 8):
              if chunk:
                  f.write(chunk)
                  f.flush()
                  os.fsync(f.fileno())
      if(isThumnail):
        img = Image.open(os.path.abspath(file_path))
        img.thumbnail((80,80),Image.ANTIALIAS)
        img.save(os.path.abspath(file_path), quality=100)

  else:  # HTTP status code 4XX/5XX
      print("Download failed: status code {}\n{}".format(r.status_code, r.text))

class XlsReaderSpider(scrapy.Spider):
  name="xlsreader"
  # Give the location of the file
  path = "G:\\mrCod3r\\projects\\15.scrapper\\imageDownloader\\foods.xlsx"
  imagesDir = r"G:\\mrCod3r\\projects\\15.scrapper\\images"
  # To open the workbook
  # workbook object is created
  wb_obj = openpyxl.load_workbook(path)
  
  # Get workbook active sheet object
  # from the active attribute
  sheet_obj = wb_obj.active
  max_col = sheet_obj.max_column
  max_row = sheet_obj.max_row
  baseUrl = 'https://Core.zirehapp.com'
  urls = []
  nameIdx = int
  isThumbnail = False
  # Loop will print all columns name
  for i in range(1, max_col + 1):
      cell_obj = sheet_obj.cell(row = 1, column = i)
      if(cell_obj.value == 'name'):
        nameIdx = cell_obj.col_idx
      if( cell_obj.value == 'thumbnail'):
        isThumbnail = True
      if(cell_obj.value == 'banner' or cell_obj.value == 'thumbnail'):
        for j in range(1,max_row+1):
          rowName = sheet_obj.cell(row = j, column = nameIdx).value.strip()
          filename = sheet_obj.cell(row = j, column = i).value.strip().split('/')[-1].replace(" ", "_")
          urlObj = {
            "name" : filename,
            "url" : sheet_obj.cell(row = j, column = i).value.strip(),
            "path":NULL,
            "isThumbnail":isThumbnail
          }
        
            
          if(len(urlObj['url']) == 0 or j == 1):
            continue
          elif(urlObj['url'].rfind("http") > -1): 
            urlObj['path'] = urlObj['url']
            urls.append(urlObj)
          else:
            urlObj['path'] = urlObj['url']
            urlObj['url'] = baseUrl+urlObj['url']
            urls.append(urlObj)
          
  for item in urls:    
    pathList = pathExtractor(item['url'])
    
    if(item['isThumbnail']== True):
      pathList.insert(0, "thumbnail")  

    Path = ''
    if(len(pathList) > 1):
      for path in pathList:
        if (len(pathList) - 1) != pathList.index(path):
          Path += '\\\\'+ path


    print(imagesDir+Path)
    download(item['url'],item['name'],imagesDir+Path,item['isThumbnail'])